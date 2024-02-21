import openpyxl
import re
import datetime

excelPageName = openpyxl.load_workbook('draftШаблон (емкости).xlsx')         # список из имен вкладок (емкости)
currentPage = excelPageName.sheetnames[-1]                      # имя последней страницы (емкости)
workSheet = excelPageName[currentPage]                          # последняя страница (емкости)

excelProduct = openpyxl.load_workbook('draftШаблон (товары).xlsx')          # список из имен вкладок (товары)
currentPageProduct = excelProduct.create_sheet(title = currentPage)         # создаем лист новый (товары)
workSheetProduct = excelProduct[currentPage]

today = datetime.datetime.today().strftime('%y%m%d')            # часть yy-mm-dd от sku

oilDict = {}                                                    # словарь для (товары) с номером sku
count_sku = 1                                                   # счетчик 001 для sku

for x in range(2, workSheet.max_row+1):
    line = (workSheet.cell(row=x, column=3).value)              # считываем 3 стобец (емкости)
    pattern = r"(^.*)\s\((\d?\.?\,?\d*)\s(\w*)\.\)$"            # паттерн для поиска RegExp (емкости)
    result = re.search(pattern, line)
    if result is None:                                          # если товар штучный или компл без литража
        workSheet.cell(row=x, column=5).value = ''
        workSheet.cell(row=x, column=8).value = ''
        oilDict[line] = today + str('{0:03}'.format(count_sku))
        sku_product = today + str('{0:03}'.format(count_sku))
        count_sku += 1
    else:
        weight = result[2].replace(',', '.')
        if result[3] == 'мл':                                   # мл как отд случай
            weight = '0.' + result[2]
        workSheet.cell(row=x, column=8).value = weight
        volume = result[2] + ' ' + result[3] + '.'
        workSheet.cell(row=x, column=5).value = volume
        if result[1] not in oilDict:                            # для выгрузки в шаблоны(товары)
            oilDict[result[1]] = today + str('{0:03}'.format(count_sku))
            sku_product = today + str('{0:03}'.format(count_sku))
            count_sku +=1
    workSheet.cell(row=x, column=1).value = sku_product
    workSheet.cell(row=x, column=4).value = 13
    workSheet.cell(row=x, column=6).value = 100
    workSheet.cell(row=x, column=7).value = 0
    workSheet.cell(row=x, column=9).value = 1

        #обработка Шаблон(товары)
y = 2
for x in oilDict:
    workSheetProduct.cell(row=y, column=1).value = oilDict[x]
    workSheetProduct.cell(row=y, column=3).value = x
    workSheetProduct.cell(row=y, column=4).value = x
    workSheetProduct.cell(row=y, column=5).value = x
    workSheetProduct.cell(row=y, column=7).value = 100
    workSheetProduct.cell(row=y, column=8).value = 0
    workSheetProduct.cell(row=y, column=10).value = 0
    workSheetProduct.cell(row=y, column=11).value = 0
    workSheetProduct.cell(row=y, column=12).value = 0
    workSheetProduct.cell(row=y, column=13).value = 'Каталог, поиск'
    y+=1

titles = ['sku', 'category_ids', 'name', 'short_description', 'description', 'product_type',
          'quantity', 'stock_status_id', 'manufacturer', 'price', 'weight', 'status', 'visibility',
          'meta_title', 'meta_description', 'meta_keyword'
]
for x in range(1, len(titles)+1):
    workSheetProduct.cell(row=1, column=x).value = titles[x-1]


excelPageName.save("C:\\Users\\Admin\\AppData\\Local\\Programs\\Python\\Python310\\venv\\draftШаблон (емкости).xlsx")
excelProduct.save("C:\\Users\\Admin\\AppData\\Local\\Programs\\Python\\Python310\\venv\\draftШаблон (товары).xlsx")

