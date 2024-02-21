import openpyxl
import re
import datetime

excelPageName = openpyxl.load_workbook('draftШаблон (емкости).xlsx')         # список из имен вкладок (емкости)
currentPage = excelPageName.sheetnames[-1]                      # имя последней страницы (емкости)
workSheet = excelPageName[currentPage]                          # последняя страница (емкости)

today = datetime.datetime.today().strftime('%y%m%d')            # часть yy-mm-dd от sku

oilDict = {}                                                    # словарь для (товары) с номером sku
count_sku = 1                                                   # счетчик 001 для sku

for x in range(2, workSheet.max_row+1):
    line = (workSheet.cell(row=x, column=3).value)              # считываем 3 стобец
    pattern = r"(^.*)\s\((\d?\.?\,?\d*)\s(\w*)\.\)$"            # паттерн для поиска RegExp
    result = re.search(pattern, line)
    if result is None:                                          # если товар штучный или компл без литража
        workSheet.cell(row=x, column=5).value = ''
        workSheet.cell(row=x, column=8).value = ''
        oilDict[line] = today + str('{0:03}'.format(count_sku))
        sku_product = today + str('{0:03}'.format(count_sku))
        count_sku += 1
    else:
        weight = result[2].replace(',', '.')
        if result[3] == 'мл':                                   # мл как отд случай
            weight = '0.' + result[2]
        workSheet.cell(row=x, column=8).value = weight
        volume = result[2] + ' ' + result[3] + '.'
        workSheet.cell(row=x, column=5).value = volume
        if result[1] not in oilDict:                            # для выгрузки в шаблоны(товары)
            oilDict[result[1]] = today + str('{0:03}'.format(count_sku))
            sku_product = today + str('{0:03}'.format(count_sku))
            count_sku +=1
    workSheet.cell(row=x, column=1).value = sku_product
    workSheet.cell(row=x, column=4).value = 13
    workSheet.cell(row=x, column=6).value = 100
    workSheet.cell(row=x, column=7).value = 0
    workSheet.cell(row=x, column=9).value = 1
        #oilDict[result[1]] = [weight, volume] #result1
print(oilDict)
excelPageName.save("C:\\Users\\Admin\\AppData\\Local\\Programs\\Python\\Python310\\venv\\draftШаблон (емкости).xlsx")

